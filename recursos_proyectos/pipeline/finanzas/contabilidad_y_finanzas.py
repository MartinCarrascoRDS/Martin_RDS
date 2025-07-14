import pandas as pd
import numpy as np
from openpyxl import load_workbook
import re

def ajustar_precios_a_margen_deseado(ruta_excel, comision = 0.15, margen_objetivo = 0.27, hoja = None, ruta_salida = None):
    """
    Ajusta el precio de venta de un conjunto de productos a un margen deseado, considerando que, dentro de los costos,
    hay una comisión por venta

    Parámetros:
    ruta_excel (str): ruta del excel que posee la información original (con márgenes distintos al objetivo).
    comisión (float): comisión por venta dentro de los costos (por defecto 15%)
    margen_objetivo (float): margen deseado (por defecto 27%)
    hoja (str|int, opcional): nombre o índice de la hoja de excel que se desea leer en caso de haber más de 1.
    ruta_salida (str, opcional): ruta donde se guardará un nuevo excel con los precios correctos para el margen deseado

    Retorna: DataFrame con precios ajustados para el margen deseado
    """

    df = pd.read_excel(ruta_excel, sheet_name = hoja)
    df.columns = df.columns.str.lower()

    columnas_requeridas = {'Producto', 'Precio de venta', 'Comisión', 'Costo total'}
    if not columnas_requeridas.issubset(df.columns):
        raise ValueError(f"""
                         Para que la función se desempeñe correctamente, debe haber una columna que indique la comisión por venta 
                         y una que indique el costo total, incluyendo comisión. En caso de haber y presentarse el error,
                         renombrarlas a {columnas_requeridas}
""")
    
    df['Costos sin comisión'] = df['Costo total'] - df['Comisión']

    denominador = 1 - margen_objetivo - comision
    if denominador <= 0:
        raise ValueError(f"""La suma entre el margen objetivo y la comisión entregada es demasiado alta como para calcular
                         un precio válido
""")
    
    df['Precio de venta ajustado'] = np.round(df['Costos sin comisión']/denominador, 2)
    df['Comisión ajustada'] = np.round(df['Precio de venta ajustado'] * comision, 2)
    df['Costo total ajustado'] = df['Costos sin comisión'] + df['Comisión ajustada']

    df['Margen ajustado'] = np.round((df['Precio de venta ajustado'] - df['Costo total ajustado'])/df['Precio de venta ajustado'], 4)

    columnas_finales = [
        'Precio de venta ajustado', 'Comisión ajustada', 'Costos sin comisión',
        'Costo total ajustado', 'Margen ajustado'
    ]

    if ruta_salida:
        df[columnas_finales].to_excel(ruta_salida, index = False)

    return df[columnas_finales]



def extraer_numero_de_paquetes(estado_str):
    match = re.search(r"Paquete de (\d+)", str(estado_str))
    return int(match.group(1)) if match else 0




def separar_sku(sku):
    if pd.isna(sku):
        return []
    partes = str(sku).upper().strip().split("/")
    resultado = []
    for parte in partes:
        parte = parte.strip()
        parte = re.sub(r"^(F-|XX-)", "", parte)
        parte = parte.replace(" /", "/")

        if " X" in parte:
            base, cantidad = parte.rsplit(" X", 1)
            try:
                cantidad_int = int(cantidad.strip())
            except:
                cantidad_int = 1

        else:
            base = parte
            cantidad_int = 1
        resultado.append((base.strip(), cantidad_int))

    return resultado



def analisis_margen(
        archivo_excel_ventas,
        archivo_excel_costos,
        nombre_hoja_ventas = None,
        nombre_hoja_costos = None,
        skiprows = 5,
        umbral_paquete = 7.5,
        dcto_cyber = True,
        fecha_inicio_cyber = "2025-06-02",
        fecha_fin_cyber = "2025-06-08"
):
    
    df = pd.read_excel(archivo_excel_ventas, skiprows = skiprows, sheet_name = nombre_hoja_ventas)

    print(f"Existen {df['# de venta'].nunique()} registros en la base de datos de ventas")

    columnas_unicas = []
    contador_estado = 0
    contador_unidades = 0
    contador_forma_entrega = 0
    columnas_eliminar = [
        "Descripción del estado", "Mes de facturación de tus cargos", "Venta por publicidad",
        "Canal de venta", "Tienda oficial", "Variante", "Tipo de publicación", "Factura adjunta",
        "Datos personales o de empresa", "Tipo y número de documento", "Dirección",
        "Tipo de contribuyente", "Actividad económica", "Comprador", "Negocio", "Cédula",
        "Domicilio", "Comuna", "Estado", "Código postal", "País", "Fecha en camino", "Fecha entregado",
        "Transportista", "Número de seguimiento", "URL de seguimiento",
        "Revisado por Mercado Libre", "Fecha de revisión",
        "Dinero a favor", "Resultado", "Destino", "Motivo del resultado",
        "Reclamo abierto", "Reclamo cerrado", "Con mediación"
    ]
    
    for col in df.columns:
        if col.startswith("Estado"):
            contador_estado += 1
            if contador_estado == 1:
                columnas_unicas.append(col)
        elif col.startswith("Unidades"):
            contador_unidades += 1
            if contador_unidades == 1:
                columnas_unicas.append(col)
        elif col.startswith("Forma de entrega"):
            contador_forma_entrega += 1
            if contador_forma_entrega == 1:
                columnas_unicas.append(col)
        elif all(not col.startswith(nombre_col) for nombre_col in columnas_eliminar):
            columnas_unicas.append(col)
    df = df[columnas_unicas].copy()

    print(df.columns)
    df['Estado'] = df['Estado'].str.strip()

    wb = load_workbook(archivo_excel_ventas, data_only = True)
    ws = wb[nombre_hoja_ventas] if nombre_hoja_ventas else wb.active
    idx_estado = list(df.columns).index("Estado")
    estados_backgrounds = []
    encabezado_fila_excel = skiprows + 1
    for row in ws.iter_rows(min_row = encabezado_fila_excel + 1, max_row = ws.max_row):
        cell = row[idx_estado]
        fill = cell.fill
        color = fill.fgColor.rgb if fill and fill.fgColor and fill.fgColor.type == "rgb" else None
        estados_backgrounds.append(color)

    paquete_indices_todos = []
    encabezados_indices = []
    i = 0
    while i < len(df):
        fondo_actual = estados_backgrounds[i]
        if fondo_actual and fondo_actual != "00000000":
            estado_valor = str(df.iloc[i]["Estado"])
            n_items = extraer_numero_de_paquetes(estado_valor)
            rango_paquete = list(range(i, i + n_items + 1))
            paquete_indices_todos.extend(rango_paquete)
            encabezados_indices.append(i)
            i += n_items + 1
        else:
            i += 1

    ingresos_total = df['Ingresos por productos (CLP)'].sum()
    ingresos_paquetes = df.iloc[encabezados_indices]['Ingresos por productos (CLP)'].sum()
    porcentaje = 100 * ingresos_paquetes / ingresos_total if ingresos_total else 0

    if porcentaje > umbral_paquete:
        print(f"Las ventas en paquete represetan un {porcentaje:.2f}% de los ingresos por venta, superando el umbral de {umbral_paquete:.2f}%")
        return None
    
    df = df.drop(index = paquete_indices_todos).reset_index(drop = True)

    print(f"{df['# de venta'].nunique()} sin considerar las ventas en paquete")

    estados_prefijos_permitidos = (
    "Acuerdas la entrega",
    "Despacharemos el paquete",
    "En camino",
    "En punto de retiro",
    "Entregado",
    "Etiqueta lista para imprimir",
    "Etiqueta para imprimir",
    "Envío reprogramado",
    "Listo para recolección",
    "Mediación finalizada. Te dimos el dinero",
    "Procesando en la bodega",
    "Venta concretada",
    "Venta entregada",
    "Venta no entregada. Te dimos el dinero"
)

    df = df[df["Estado"].apply(lambda x: any(x.startswith(p) for p in estados_prefijos_permitidos))].reset_index(drop=True)

    meses_es = {
        'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
        'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
        'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12'
    }
    def convertir_fechas(fecha_str):
        if pd.isna(fecha_str):
            return pd.NaT
        try:
            fecha_str = str(fecha_str).lower()
            fecha_str = re.sub(r"\s*hs\.?", "", fecha_str)
            partes = fecha_str.split(' de ')
            if len(partes) < 3:
                return pd.NaT
            dia = partes[0].strip().zfill(2)
            mes = meses_es.get(partes[1].strip(), '01')
            año_hora = partes[2].strip()
            año = año_hora.split()[0]
            return pd.to_datetime(f"{año}/{mes}/{dia}", format="%Y/%m/%d")
        except Exception:
            return pd.NaT

    if "Fecha de venta" in df.columns:
        df["Fecha de venta"] = df["Fecha de venta"].apply(convertir_fechas)

    print(f"De todos los registros, {df['# de venta'].nunique()} son ventas")

    columnas_numericas = [
        "Unidades", "Ingresos por productos (CLP)", "Cargo por venta e impuestos (CLP)",
        "Ingresos por envío (CLP)", "Costos de envío (CLP)"
    ]

    for col in columnas_numericas:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors = "coerce")
    for col in ["Ingresos por envío (CLP)", "Costos de envío (CLP)"]:
        df[col] = df[col].fillna(0)
    for col in [
        "Ingresos por productos (CLP)", "Cargo por venta e impuestos (CLP)",
        "Ingresos por envío (CLP)", "Costos de envío (CLP)"
    ]:
        df[f"{col} Neto"] = df[col] / 1.19
    
    registros = []
    for idx, row in df.iterrows():
        skus = separar_sku(row['SKU'])
        for sku, cantidad in skus:
            nuevo = row.copy()
            nuevo["Proveedor"] = sku.split("-")[0].strip()
            nuevo["SKU limpio"] = sku
            nuevo["Cantidad SKU"] = cantidad
            registros.append(nuevo)
    df_expandido = pd.DataFrame(registros)

    costos = pd.read_excel(archivo_excel_costos, sheet_name = nombre_hoja_costos, usecols = ["SKU", "PRECIO"])
    costos["SKU"] = costos["SKU"].str.upper().str.strip()
    df_expandido["SKU limpio"] = df_expandido["SKU limpio"].str.upper().str.strip()
    costos_antes = costos.shape[0]
    costos = costos.drop_duplicates(subset = ['SKU'], keep = 'first')
    costos_despues = costos.shape[0]
    print(f"Se eliminaron {costos_antes - costos_despues} SKU repetidos de la base de datos de costos")
    df_merged = df_expandido.merge(costos[['SKU', 'PRECIO']], how = "left", left_on = "SKU limpio", right_on = "SKU")
    df_merged["PRECIO FALTANTE"] = df_merged['PRECIO'].isna()

    ventas_totales = df_merged['# de venta'].nunique()
    ventas_con_faltantes = df_merged.groupby('# de venta')["PRECIO FALTANTE"].any()
    ventas_buenas = ventas_con_faltantes[~ventas_con_faltantes].index

    df_merged = df_merged[df_merged['# de venta'].isin(ventas_buenas)].copy()
    ventas_restantes = df_merged['# de venta'].nunique()

    print(f"Se eliminaron {ventas_totales - ventas_restantes} ventas completas por tener algún SKU sin precio.")

    df_merged['Costo unitario'] = df_merged['PRECIO']

    if dcto_cyber:
        fecha_inicio_cyber = pd.to_datetime(fecha_inicio_cyber)
        fecha_fin_cyber = pd.to_datetime(fecha_fin_cyber)
        condiciones_cyber = df_merged["Fecha de venta"].between(fecha_inicio_cyber, fecha_fin_cyber)
        descuentos = df_merged["Proveedor"].map({
            "MA": 0.95, "RX": 0.9, "CR": 0.97, "AL": 0.96, "NC": 0.96 
        })
        df_merged["Costo unitario con descuento"] = df_merged["Costo unitario"]
        df_merged.loc[condiciones_cyber, "Costo unitario con descuento"] = (
            df_merged.loc[condiciones_cyber, "Costo unitario"] * descuentos
        )

    df_merged["Costo producto"] = df_merged["Costo unitario con descuento"] * df_merged["Cantidad SKU"]

    agrupado = df_merged.groupby("# de venta").agg({
        "Fecha de venta": "first",
        "Estado": "first",
        "Unidades": "first",
        "Forma de entrega": "first",
        "Ingresos por productos (CLP) Neto": "first",
        "Cargo por venta e impuestos (CLP) Neto": "first",
        "Ingresos por envío (CLP) Neto": "first",
        "Costos de envío (CLP) Neto": "first",
        "Costo producto": "sum"
    }).reset_index()

    agrupado["Costo producto total"] = agrupado["Costo producto"] * agrupado["Unidades"]
    agrupado["Costo total"] = (
        agrupado["Costo producto total"] -
        agrupado["Cargo por venta e impuestos (CLP) Neto"] -
        agrupado["Ingresos por envío (CLP) Neto"] -
        agrupado["Costos de envío (CLP) Neto"]
    )

    agrupado["Margen"] = (
        (agrupado["Ingresos por productos (CLP) Neto"] - agrupado["Costo total"]) / agrupado["Ingresos por productos (CLP) Neto"]
    )

    agrupado['# de venta'] = agrupado['# de venta'].astype(str)

    return agrupado