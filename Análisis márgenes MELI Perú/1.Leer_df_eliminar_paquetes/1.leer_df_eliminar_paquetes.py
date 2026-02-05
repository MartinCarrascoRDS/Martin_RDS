import os
import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from pipeline.procesamiento.funciones_para_analisis_margen import extraer_año_desde_fecha, extraer_numero_de_paquetes, convertir_fechas

# -------------------------------------------------------------------------
# Carga de archivo
# -------------------------------------------------------------------------

cuenta_meli = input('Indique la cuenta de Mercado Libre a la que corresponde este análisis (ejemplo: BLACKPARTS): ')
fecha = input('Indique la fecha del análisis (ejemplo: JUNIO 2025): ')
año = extraer_año_desde_fecha(fecha)
print(f'El año de análisis es {año}')

archivo_venta = f'/Users/martincarrasco/Desktop/Martín_Carrasco/Reportes/{año}/Cuentas RDS/{cuenta_meli}/VENTAS {cuenta_meli} {fecha}.xlsx'
hoja_venta = 'Ventas PE'
skiprows = 5

df = pd.read_excel(
    archivo_venta,
    sheet_name=hoja_venta,
    skiprows=skiprows,
    dtype={"# de venta": str}
)

print(f"Existen {df['# de venta'].nunique()} registros en la base de datos de ventas")

# Formato fecha
if 'Fecha de venta' in df.columns:
    df['Fecha de venta'] = df['Fecha de venta'].apply(convertir_fechas)

# FFill en "Forma de entrega"
df["Forma de entrega"] = df["Forma de entrega"].replace(r'^\s*$', np.nan, regex=True)
df["Forma de entrega"] = df["Forma de entrega"].fillna(method='ffill')

df['Tipo de venta'] = 'Unitaria'

df['Ingresos por productos (PEN)'] = df['Ingresos por productos (PEN)'].fillna(
    df['Unidades'] * df['Precio unitario de venta de la publicación (PEN)']
)

# -------------------------------------------------------------------------
# Detectar paquetes usando el color de fondo
# -------------------------------------------------------------------------

wb = load_workbook(archivo_venta, data_only=True)
ws = wb[hoja_venta]

idx_estado = list(df.columns).index('Estado')
estados_backrounds = []

inicio_datos_excel = skiprows + 1

for row in ws.iter_rows(min_row=inicio_datos_excel + 1, max_row=ws.max_row):
    cell = row[idx_estado]
    fill = cell.fill
    color = fill.fgColor.rgb if fill and fill.fgColor and fill.fgColor.type == "rgb" else None
    estados_backrounds.append(color)

# -------------------------------------------------------------------------
# Identificar bloques de paquete
# -------------------------------------------------------------------------

paquete_encabezados = []
paquete_indices = []  # TODAS las filas del rango paquete

i = 0
while i < len(df):
    fondo = estados_backrounds[i]
    estado_txt = str(df.iloc[i]['Estado'])

    if fondo and fondo != "00000000" and "Paquete de" in estado_txt:
        n_items = extraer_numero_de_paquetes(estado_txt)
        rango = list(range(i, i + n_items + 1))
        paquete_indices.extend(rango)
        paquete_encabezados.append(i)
        i += n_items + 1
    else:
        i += 1

print(f"Hubo {len(paquete_encabezados)} ventas en paquete")

df.loc[paquete_indices, 'Tipo de venta'] = 'Paquete'

# -------------------------------------------------------------------------
# AGREGAR COLUMNA No. Paquete
# -------------------------------------------------------------------------

df['No. Paquete'] = "Unitaria"

for idx_head in paquete_encabezados:
    estado_txt = df.at[idx_head, 'Estado']

    m = re.search(r"Paquete de (\d+)", estado_txt)
    if not m:
        continue

    n_items = int(m.group(1))
    no_venta = df.at[idx_head, "# de venta"]

    for i in range(1, n_items + 1):
        fila = idx_head + i
        if fila in df.index:
            df.at[fila, "No. Paquete"] = no_venta

df['No. Paquete'] = df['No. Paquete'].astype(str)

# -------------------------------------------------------------------------
# PRORRATEO COMPLETO A CADA PRODUCTO DEL PAQUETE
# -------------------------------------------------------------------------

df_prorrateado = []

for idx_head in paquete_encabezados:

    fila_head = df.loc[idx_head]
    n_items = extraer_numero_de_paquetes(fila_head["Estado"])
    filas_items = df.loc[idx_head+1 : idx_head+n_items].copy()

    if filas_items.empty:
        continue

    total_ingresos = filas_items["Unidades"] * filas_items["Precio unitario de venta de la publicación (PEN)"]
    suma_ingresos = total_ingresos.sum()

    if suma_ingresos == 0:
        continue

    filas_items["% participación"] = total_ingresos / suma_ingresos

    costos_prorratear = [
        "Cargo por venta e impuestos (PEN)",
        "Ingresos por envío (PEN)",
        "Costos de envío (PEN)"
    ]

    for col in costos_prorratear:
        filas_items[col] = filas_items["% participación"] * fila_head[col]

    filas_items = filas_items.drop(columns=["% participación"])

    df_prorrateado.append(filas_items)

df_prorrateado = pd.concat(df_prorrateado, ignore_index=True) if df_prorrateado else pd.DataFrame()

# -------------------------------------------------------------------------
# Remover filas originales de paquete y reemplazar por prorrateo
# -------------------------------------------------------------------------

df_clean = df.drop(index=paquete_indices).reset_index(drop=True)
df_final = pd.concat([df_clean, df_prorrateado], ignore_index=True)

indices_cambio = df_final.index[df_final['Estado'] == 'Venta con solicitud de cambio']
for idx in indices_cambio:
    df_final.at[idx, 'Tipo de venta'] = 'Cambio'
    if idx + 1 in df_final.index:
        df_final.at[idx + 1, 'Tipo de venta'] = 'Cambio'
    if idx + 2 in df_final.index:
        df_final.at[idx + 2, 'Tipo de venta'] = 'Cambio'

df_final = df_final[df_final['Tipo de venta'] != 'Cambio'].copy()

print(f"Registros finales luego de integrar paquetes: {df_final.shape[0]}")

# -------------------------------------------------------------------------
# Exportar resultados
# -------------------------------------------------------------------------

output_folder = f'/Users/martincarrasco/Desktop/Martín_Carrasco/Análisis márgenes MELI Perú/1.Leer_df_eliminar_paquetes/{año}/{fecha}'
os.makedirs(output_folder, exist_ok=True)
output_path = f'{output_folder}/{fecha} {cuenta_meli} TOTAL VENTAS.xlsx'
df_final.to_excel(output_path, index=False)

print("Proceso completado con éxito.")