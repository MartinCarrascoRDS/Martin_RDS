"""
Paso 1:
Leer un archivo excel donde se encuentren las ventas como dataframe, y eliminar las ventas en paquete.
Calcular las ventas totales, las ventas en paquete, y el porcentaje que estas últimas representan.
"""

import pandas as pd
from openpyxl import load_workbook
import re

def extraer_numero_de_paquetes(estado_str):
    match = re.search(r"Paquete de (\d+)", str(estado_str))
    return int(match.group(1)) if match else 0

archivo_venta = '/Users/martincarrasco/Desktop/Martín_Carrasco/Reportes/2025/Cuentas RDS/Blackpartscl/VENTAS BLACKPARTS JUNIO 2025.xlsx'
hoja_venta = 'Ventas CL'
skiprows = 5
df = pd.read_excel(archivo_venta, sheet_name = hoja_venta, skiprows = skiprows)
print(f"Existen {df['# de venta'].nunique()} registros en la base de datos de ventas")

wb = load_workbook(archivo_venta, data_only = True)
ws = wb[hoja_venta] if hoja_venta else wb.active
idx_estado = list(df.columns).index('Estado')
estados_backrounds = []
encabezado_fila_excel = skiprows + 1
for row in ws.iter_rows(min_row = encabezado_fila_excel + 1, max_row = ws.max_row):
    cell = row[idx_estado]
    fill = cell.fill
    color = fill.fgColor.rgb if fill and fill.fgColor and fill.fgColor.type == "rgb" else None
    estados_backrounds.append(color)

paquete_indices_todos = []
encabezados_indices = []
i = 0
while i < len(df):
    fondo_actual = estados_backrounds[i]
    if fondo_actual and fondo_actual != "00000000":
        estado_valor = str(df.iloc[i]['Estado'])
        n_items = extraer_numero_de_paquetes(estado_valor)
        rango_paquete = list(range(i, i + n_items + 1))
        paquete_indices_todos.extend(rango_paquete)
        encabezados_indices.append(i)
        i += n_items + 1
    else:
        i += 1

ingreso_total = df['Ingresos por productos (CLP)'].sum()
ingreso_paquetes = df.iloc[encabezados_indices]['Ingresos por productos (CLP)'].sum()
porcentaje = 100 * ingreso_paquetes / ingreso_total if ingreso_total else 0

print(f'Las ventas en paquete representan el {porcentaje:.2f}% de las ventas')

df = df.drop(index = paquete_indices_todos).reset_index(drop = True)

print(f"Existen {df['# de venta'].nunique()} registros en la base de datos de ventas sin contar las ventas en paquete")

df['# de venta'] = df['# de venta'].astype(str)

output_path = '/Users/martincarrasco/Desktop/Martín_Carrasco/Testeo Reportes/Análisis márgenes/1.Leer_df_eliminar_paquetes/Paso1_listo.xlsx'
df.to_excel(output_path, index=False)
