"""
Paso 1:
Leer un archivo excel donde se encuentren las ventas como dataframe, y eliminar las ventas en paquete.
Calcular las ventas totales, las ventas en paquete, y el porcentaje que estas últimas representan.
Aislar ventas en paquete, y ventas totales (considerando paquete)
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import re
import os
from pipeline.procesamiento.procesamiento_bases import convertir_fechas

def extraer_numero_de_paquetes(estado_str):
    match = re.search(r"Paquete de (\d+)", str(estado_str))
    return int(match.group(1)) if match else 0

año = 2024 # RECORDAR CAMBIAR EL AÑO PARA GENERAR NUEVAS CARPETAS

cuenta_meli = input('Indique la cuenta de Mercado Libre a la que corresponde este análisis (ejemplo: BLACKPARTS): ')
fecha = input('Indique la fecha del análisis (ejemplo: JUNIO 2025): ')

archivo_venta = f'/Users/martincarrasco/Desktop/Martín_Carrasco/Reportes/{año}/Cuentas RDS/{cuenta_meli}/VENTAS {cuenta_meli} {fecha}.xlsx'
hoja_venta = 'Ventas CL'
skiprows = 5
df = pd.read_excel(archivo_venta, sheet_name = hoja_venta, skiprows = skiprows, dtype = {"# de venta": str})
print(f"Existen {df['# de venta'].nunique()} registros en la base de datos de ventas")

if 'Fecha de venta' in df.columns:
    df['Fecha de venta'] = df['Fecha de venta'].apply(convertir_fechas)

df['Forma de entrega'] = df['Forma de entrega'].replace(r'^\s*$', np.nan, regex=True)
df["Forma de entrega"] = df["Forma de entrega"].fillna(method = 'ffill')

df_ingresos = df[["# de venta", "Fecha de venta", "Estado", "Unidades", "SKU", "# de publicación", "Título de la publicación", "Precio unitario de venta de la publicación (CLP)", "Forma de entrega"]]
df_ingresos["Cuenta Meli"] = cuenta_meli

output_folder = f'/Users/martincarrasco/Desktop/Martín_Carrasco/Análisis márgenes/1.Leer_df_eliminar_paquetes/{año}/{fecha}'
os.makedirs(output_folder, exist_ok = True)

output_path_ingresos = f'{output_folder}/Paso1_totales_{cuenta_meli}_{fecha}_listo.xlsx'
df_ingresos.to_excel(output_path_ingresos, index = False)

wb = load_workbook(archivo_venta, data_only = True)
ws = wb[hoja_venta] if hoja_venta else wb.active
idx_estado = list(df.columns).index('Estado')
estados_backrounds = []
encabezado_fila_excel = skiprows + 1
for row in ws.iter_rows(min_row = encabezado_fila_excel + 1, max_row = ws.max_row):
    cell = row[idx_estado]
    fill = cell.fill
    color = fill.fgColor.rgb if fill and fill.fgColor and fill.fgColor.type == "rgb" else None
    estados_backrounds.append(color)

paquete_indices_todos = []
encabezados_indices = []
i = 0
while i < len(df):
    fondo_actual = estados_backrounds[i]
    estado_valor = str(df.iloc[i]['Estado'])
    if fondo_actual and fondo_actual != "00000000" and "Paquete de" in estado_valor:
        n_items = extraer_numero_de_paquetes(estado_valor)
        rango_paquete = list(range(i, i + n_items + 1))
        paquete_indices_todos.extend(rango_paquete)
        encabezados_indices.append(i)
        i += n_items + 1
    else:
        i += 1

print(f'Hubo {len(encabezados_indices)} ventas en paquete')

df_encabezados = pd.read_excel(archivo_venta, sheet_name = hoja_venta, skiprows = skiprows, dtype = {"# de venta": str})
df_encabezados = df_encabezados.iloc[encabezados_indices].copy()
df_paquetes = df_encabezados[["# de venta", "Fecha de venta", "Ingresos por productos (CLP)"]]
df_paquetes['Cuenta Meli'] = cuenta_meli

print(encabezados_indices)

ingreso_total = df['Ingresos por productos (CLP)'].sum()
ingreso_paquetes = df.iloc[encabezados_indices]['Ingresos por productos (CLP)'].sum()
porcentaje = 100 * ingreso_paquetes / ingreso_total if ingreso_total else 0

print(f'Las ventas en paquete representan el {porcentaje:.2f}% de las ventas')

df = df.drop(index = paquete_indices_todos).reset_index(drop = True)

df['# de venta'] = df['# de venta'].astype(str)

df = df[df["Ingresos por productos (CLP)"].notna()]

print(f"Existen {df['# de venta'].nunique()} registros en la base de datos de ventas sin contar las ventas en paquete, y con un ingreso válido")

output_path = '/Users/martincarrasco/Desktop/Martín_Carrasco/Análisis márgenes/1.Leer_df_eliminar_paquetes/Paso1_listo.xlsx'
df.to_excel(output_path, index=False)

output_path_paquetes = f'{output_folder}/Paso1_ventas_paquete_{cuenta_meli}_{fecha}.xlsx'
df_paquetes.to_excel(output_path_paquetes, index = False)